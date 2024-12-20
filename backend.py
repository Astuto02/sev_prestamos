
from PySide6.QtWidgets import (QWidget, QFrame,QLabel, QHBoxLayout,QVBoxLayout, QPushButton, QTableWidget, QAbstractItemView, QLineEdit,
                                QGridLayout, QDialog,QFileDialog, QComboBox, QTextEdit,QTableWidgetItem,QDateEdit, QHeaderView,
                                QMessageBox,QTableView,QProgressBar,QCompleter)
from PySide6.QtGui import QIcon, QPixmap, QPainter
from PySide6.QtCore import  QModelIndex, QPersistentModelIndex, QSize,Qt,QDateTime, QDate, Signal,QAbstractTableModel,Slot,Signal,QBasicTimer
from datetime import datetime
from PySide6.QtCharts import QChart,QChartView,QBarSeries,QBarCategoryAxis,QBarSet,QValueAxis
from PySide6.QtPrintSupport import QPrinter,QPrintDialog, QPrintPreviewDialog
from PySide6.QtSql import QSqlDatabase,QSqlQueryModel



from PySide6.QtPrintSupport import QPrinter,QPrintDialog, QPrintPreviewDialog
import os
from bd_prestamos import Comunicacion,Conexion



class Tabla(QTableWidget):
    #row_selected = Signal(list)
    def __init__(self,columns_label):
        super().__init__()
        self.columns_label = columns_label
        self.datos  = []
        
        self.verticalHeader().hide()
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.horizontalHeader().setStretchLastSection(True)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setDragDropOverwriteMode(False)
        self.setTextElideMode(Qt.ElideRight)
        self.setWordWrap(False)
        self.setSortingEnabled(False)
        self.setColumnCount(len(columns_label))
        self.horizontalHeader().setDefaultAlignment(Qt.Alignment.AlignCenter)
        self.setAlternatingRowColors(True)
        self.setHorizontalHeaderLabels(columns_label)
        self.setAlternatingRowColors(True)
        #self.setSelectionModel(QAbstractItemView.SingleSelection)
        #self.setModel(self._create_model())
        #self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.clearContents()

        
    
        
class SelecRow:
    '''Selecionar un elemento dentro de una tabla
    se le proporciona la tabla widget'''
    def __init__(self,tabla):
        super().__init__()

        self.tabla = tabla
        self.valor_fila_seleccionada = None
        #self.posicion = posicion
        
        
    def obten_valor(self, valor:int):
        '''recibe el valor de la posicion que se necesita simpre un numero entero'''
        fila = self.tabla.selectedItems()
        indice= fila[0].row()
        if fila:
            fila=[dato.text() for dato in fila]
            code = ([self.tabla.item(indice,valor).text()])
            return(code[0])
        if indice ==0:
            #QMessageBox(os.path.join('img/img_petitt.ico'),'Atencion','Porfavor seleccione un elemento')
            QMessageBox('Atencion','Porfavor seleccione un elemento')

class BarraProgreso(QProgressBar):
    def __init__(self,status,btn_conexion,conexion):
        super().__init__()

        self.status = status
        self.btn_conexion = btn_conexion
        self.connected = conexion

        self.timer = QBasicTimer()
        self.step = 0


    def timerEvent(self, e):
        if self.step >= 51 and not self.connected:
            self.timer.stop()
            self.status.setText('Connection Failed')
            self.btn_conexion.setText('Retry')
            return
        if self.step >= 100 and self.connected:
            self.timer.stop()
            self.status.setText('Connection Succesful')
            return
        self.step +=1
        self.setValue(self.step)
        

class OperacionesMath:
    def __init__(self,numero1=0,numero2=0,numero3= 0):
        super().__init__()

        self.numero1 = numero1
        self.numero2 = numero2
        self.numero3 = numero3

    def suma(self):
        return ((self.numero1 + self.numero2) + self.numero3)
    
    def resta(self):
        return (self.numero1 - self.numero2)
    
    def mulipli(self):
        return (self.numero1 * self.numero2)
    
    def division(self):
        return (self.numero1 / self.numero2)
        
class BuscarElementos():
    def __init__(self,categoria=[]):
        self.categoria = categoria

        
        self.l_info_busque = QHBoxLayout()

        self.search_data = QLineEdit(objectName='buscar')
        self.search_data.setFixedWidth(500)

        self.category_data = QComboBox(objectName='buscar')
        self.category_data.setFixedWidth(100)
        #self.category_data=[self.categoria]
        for elemento in self.categoria:
            self.category_data.addItem(elemento.capitalize())

        self.search_btn_data = QPushButton()
        self.search_btn_data.setFlat(True)
        self.search_btn_data.setIcon(QIcon(QPixmap(os.path.join("icons/busqueda.png"))))
        self.search_btn_data.setIconSize(QSize(16,16))

        self.l_info_busque.addWidget(self.category_data)
        self.l_info_busque.addWidget(self.search_data)
        self.l_info_busque.addWidget(self.search_btn_data)






class MostrarDatos_Resumen_peaje():
    '''Recibe el nombre del objeto tabla widget'''
    def __init__(self,tabla= None):
        super().__init__()

        self.tabla = tabla

        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()


    def addDataTable(self,fecha):
        '''mostrar datos dentro de la tabla, solo para Resumen Peaje'''
        self.registro = self.datos_productos.suma_resumen_peaje(fecha)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1

class MostrarDatos_Resumen_Barrera():
    '''Recibe el nombre del objeto tabla widget'''
    def __init__(self,tabla= None):
        super().__init__()

        self.tabla = tabla

        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()


    def addDataTable(self,fecha):
        '''mostrar datos dentro de la tabla, solo para Resumen Peaje'''
        self.registro = self.datos_productos.suma_resumen_barrera(fecha)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1
class MostrarDatos_in():
    '''Recibe el nombre del objeto tabla widget'''
    def __init__(self,tabla= None,combo = None, lineEdit=None):
        super().__init__()

        self.tabla = tabla
        self.combo = combo
        self.lineEdit = lineEdit

        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()

    def addDataTable(self,campos,tabla):
        '''mostrar datos dentro de la tabla, recibe los campos de la tabla segun la necesidad y el nombre de la tabla bd'''
        self.registro = self.datos_productos.mostrar_datos_tabla(campos,tabla)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1

    def addDataTable_filter_max(self,columnas,tabla_bd,columna,celda,columna2,celda2,columna3,celda3):
        '''mostrar datos dentro de la tabla, recibe los campos de la tabla segun la necesidad y el nombre de la tabla bd'''
        self.registro = self.datos_productos.filtro_condicion_date(columnas,tabla_bd,columna,celda,columna2,celda2,columna3,celda3)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1

    def addDataTableCondicion(self,columnas=None,tabla=None,columna=None,celda=None):
        '''una busqueda normalita, la lista de las columnas que quieres obtener,
          nombre de la tabla, nombre de la columna de la base de datos que tener en cuenta
            y el nombre de la variable con la que comparar'''
        self.registro = self.datos_productos.filtro_condicion(columnas, tabla,columna,celda)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1
    
    def addDataCombo(self,columnas, tabla,columna,celda):
        self.registro = self.datos_productos.filtro_condicion(columnas, tabla,columna,celda)
        if self.combo != None:
            for j in self.registro:
                self.lista.append(j[0])
                self.combo.addItem(j[0])

    def addAutofill(self,columnas, tabla,columna,celda):
        self.registro = self.datos_productos.filtro_condicion(columnas, tabla,columna,celda)
        return self.registro
        


        
class MostrarData_in():
    '''Recibe el nombre del objeto tabla widget'''
    def __init__(self,tabla= None):
        super().__init__()

        self.tabla = tabla

        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Conexion()


    def addDataTable(self,campos,tabla):
        '''mostrar datos dentro de la tabla, recibe los campos de la tabla segun la necesidad y el nombre de la tabla bd'''
        self.registro = self.datos_productos.mostrar_datos_tabla(campos,tabla)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1
class MostrarBusquedaProductos():
    '''en principio recibe el nombre de la tabla'''
    def __init__(self,tabla= None,combo = None):
        super().__init__()

        self.tabla = tabla
        self.combo = combo
        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()


    def addResultSearch(self,columnas,tabla,columna,celda,columna2,celda2,columna3,celda3):
        '''recibe los datos de estas variables para un filtro algo complejo;
            columnas: de la base de datos que quieres obtener
            tabla: el nombre de la tabla de la base de datos
            columna: columna especifica de la base de datos que tener en cuenta
            celda: nombre de la variable de la aplicacion cn la que comparar
            columna2: nombre de la segunda columna de la base de datos que tener en cuenta
            celda2: nombre de la variable de la aplicacion con la que comparar

        '''
        self.registro = self.datos_productos.filtro_condicion_date(columnas,tabla,columna,celda,columna2,celda2,columna3,celda3)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1
        

class MostrarDatosBy():
    def __init__(self,tabla= None):
        super().__init__()

        self.tabla = tabla

        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()


    def addDataTable(self,campos,tablabd,agrupar):
        '''recibe los campos de filtro el nombre de la tabla y el nombre de la xolumna con la que agrupa'''
        self.registro = self.datos_productos.mostrar_datos_tabla_comb(campos,tablabd,agrupar)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1
class MostrarDatosFiltro():
    def __init__(self,tabla= None,combo = None):
        super().__init__()

        self.tabla = tabla
        self.combo = combo
        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()


    def addDataTableCondicion(self,columnas=None,tabla=None,columna=None,celda=None):
        '''una busqueda normalita, la lista de las columnas que quieres obtener,
          nombre de la tabla, nombre de la columna de la base de datos que tener en cuenta
            y el nombre de la variable con la que comparar'''
        self.registro = self.datos_productos.filtro_condicion(columnas, tabla,columna,celda)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1

    def addDataCombo(self,columnas, tabla,columna,celda):
        self.registro = self.datos_productos.filtro_condicion(columnas, tabla,columna,celda)
        if self.combo != None:
            for j in self.registro:
                self.lista.append(j[0])
                self.combo.addItem(j[0])

class MostrarResumenFiltro():
    def __init__(self,tabla= None,combo = None):
        super().__init__()

        self.tabla = tabla
        self.combo = combo
        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()


    def addDataTableCondicion(self,columnas=None,tabla=None,columna=None,fecha=None,peajes=None):
        '''una busqueda normalita, la lista de las columnas que quieres obtener,
          nombre de la tabla, nombre de la columna de la base de datos que tener en cuenta
            y el nombre de la variable con la que comparar'''
        self.registro = self.datos_productos.resumen_peaje(columnas, tabla,columna,fecha,peajes)
        if self.tabla != None:
            i = len(self.registro)
            self.tabla.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    self.tabla.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1

    def addDataCombo(self,columnas, tabla,columna,celda):
        self.registro = self.datos_productos.filtro_condicion(columnas, tabla,columna,celda)
        if self.combo != None:
            for j in self.registro:
                self.lista.append(j[0])
                self.combo.addItem(j[0])

                    
class MostrarDatos():
    def __init__(self, columnas,tabla):
        self.columnas = columnas
        self.tabla = tabla
        self.lista = list()
        self.diccionario= dict()
        self.datos_productos = Comunicacion()

        self.registro = self.datos_productos.mostrar_datos(self.columnas, self.tabla)
    def addDataTable(self,tabla_pro):
        if tabla_pro != None:
            i = len(self.registro)
            tabla_pro.setRowCount(i)
            tablerow = 0
            for row in range(len(self.registro)):
                for column in range(len(self.registro[0])):
                    tabla_pro.setItem(row,column,QTableWidgetItem(str(self.registro[row][column])))
                    tablerow +=1
    

    def addDataCombo(self,combo):
        if combo != None:
            for j in self.registro:
                self.lista.append(j[0])
                combo.addItem(j[0])

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import openpyxl

class EditarExcel():
    def __init__(self, file_path):
        self.file_path=file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    def actualizar_celda(self,cell,valor):
        self.ws[cell]=valor

    def add_imagen(self,ruta_imagen,celda):
        imagen = Image(ruta_imagen)
        self.ws.add_image(imagen,celda)
        

    def add_tupla_column(self, data_tupla,start_row=1, start_col=1):
        for row_idx, row_data in enumerate( data_tupla, start=start_row):
            for col_idx,cell_value in enumerate(row_data,start=start_col):
                self.ws.cell(row =row_idx, column=col_idx, value=cell_value)
    
    def limpiar_filas(self,start_cell, end_cell):
        start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
        end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)
        start_col = openpyxl.utils.column_index_from_string(start_col)
        end_col=openpyxl.utils.column_index_from_string(end_col)

        for row in  range (start_row,end_row +1): 
            for col in range(start_col, end_col +1):
                self.ws.cell(row=row, column=col).value=None

        

    def save(self):
        self.wb.save(self.file_path)


class BarChar(QChartView):
    #categorias = ['enero','Febrero','Marzo']
    

    def __init__(self, titulo =None,categoria=[]):
        super().__init__()

        self.numeros = {
        'Asonga':[100,200,500],
        'Bome':[500,30,50],
        'Bolondo':[450,330,640],
        'Aman':[900,340,760],
        'Adjap':[610,340,760],
        'Bindung':[760,900,210],
        'Bidiba':[850,1000,810],
        'Nkoatoma':[340,700,510]
            }

        series = QBarSeries()
        #self.categoria = categoria
        #series = QStackedBarSeries

        for key, values in self.numeros.items():
            barset = QBarSet(key)
            for value in values:
                barset.append(value)
            series.append(barset)

        chart = QChart()
        chart.addSeries(series)
        chart.setTitle(titulo) 
        chart.setAnimationOptions(QChart.SeriesAnimations)

        axis_x = QBarCategoryAxis()
        axis_x.append(categoria)
        chart.addAxis(axis_x, Qt.AlignBottom)
        series.attachAxis(axis_x)
        

        axis_y = QValueAxis()
        chart.addAxis(axis_y, Qt.AlignLeft)
        series.attachAxis(axis_y)

        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)

        self.setRenderHint(QPainter.Antialiasing)
        self.setChart(chart)



import win32api
import win32print
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

class Impresion():
    def __init__(self,file_path):
        self.file_path = file_path
        

    def imprimir(self):
        self.file = f'{self.file_path}.xlsx'
        win32api.ShellExecute(0,'print',self.file,win32print.GetDefaultPrinter(),'.',0)
    
    def guardar_pdf(self,excel,pdf):
        archivo_pdf =f'{self.file_path}.pdf'
        ruta = os.path.join(os.path.expanduser('~'),'Desktop')
        new_folder = 'Archivos_PDF'
        new_path = os.path.join(ruta,new_folder)
        if not os.path.exists(new_path):
            os.makedirs(new_path)
        wb = openpyxl.load_workbook(self.file)
        sheet = wb.active

        c =canvas.Canvas(new_path,pagesize=letter)
        width, height = letter
        x_offset = 50
        y_offset = height -50
        row_height = 20
        for row in sheet.iter_rows(values_only = True):
            for col_index, cell_value in enumerate(row):
                x_position = x_offset + col_index * 100
                c.drawString(x_position,y_offset,str(cell_value))
            y_offset -= row_height
        c.save()
